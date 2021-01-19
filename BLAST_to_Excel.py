#import os
from os import path
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell

class Sample:
    def __init__(self, name, ct_positive):
        self.name = name
        self.ct_positive = ct_positive
        self.num_species_kraken = 0
        self.reads = []

    def addRead(self, n, blast_hits, length):
        self.reads.append(Read(n, blast_hits, length))

    def setFraction(self, read, hit, frac):
        for r in self.reads:
            if r.n == read:
                r.setFraction2(hit, frac)

    def setPercent(self, read, hit, percent):
        for r in self.reads:
            if r.n == read:
                r.setPercent2(hit, percent)

    def set_cds(self, read, hit, start, end):
        for r in self.reads:
            if r.n == read:
                r.set_cds2(hit, start, end)

class Read:
    def __init__(self, n, blast_hits, length):
        self.n = n
        self.length = length
        self.blast_hits = []
        for hit in blast_hits:
            line = hit[0:68]
            score = hit[70:74]
            e_val = hit[78:83]
            self.blast_hits.append(BlastHit(line, score, e_val))

    def setFraction2(self, hit, frac):
        self.blast_hits[hit-1].fraction = frac

    def setPercent2(self, hit, percent):
        self.blast_hits[hit-1].percent = percent

    def set_cds2(self, hit, start, end):
        self.blast_hits[hit-1].cds_start = start
        self.blast_hits[hit-1].cds_end = end

    def is_species(self, species):
        for hit in self.blast_hits:
            if species in hit.line:
                return True
        return False

class BlastHit:
    def __init__(self, line, score, e_val):
        self.line = line
        self.cds_start = ""
        self.cds_end = ""
        self.score = score
        self.e_val = e_val
        self.fraction = ""
        self.percent = ""

"""Searches through kraken reports to generate column C of the Excel file."""
def count_kraken(species, samples):
    print("Counting total number of reads for " + species)
    for s in samples:
        f = open("/home/sbomman/Kraken2/output/kraken_report_clean_" + s.name + ".txt", "r")
        line = f.readline()
        while line != "":
            if species in line:
                words = line.split()
                s.num_species_kraken = words[1]
                break
            line = f.readline()
    print("Finished counting total number of reads for " + species)

"""Calculates column D of the Excel file."""
def count_matches(species, sample):
    count = 0
    for read in sample.reads:
        if read.is_species(species):
            count += 1
    return count

"""Calculates all the other columns and stores the data in the sample objects."""
def other_columns(species, samples, blast_dir):
    print("Starting other columns for " + species)
    for s in samples:
        read_num = 0
        filename = blast_dir + s.name + "_blasted.txt"
        if path.exists(filename):
            print("Sample " + s.name + ": " + str(s.num_species_kraken) + " reads")
            f = open(filename, "r")
            line = f.readline()
            while line != "":
                if "Query=" in line:
                    read_num += 1
                    print("Read " + str(read_num))
                    f.readline()
                    line = f.readline()
                    length = line[line.index('=')+1 :]
                    f.readline()
                    f.readline()
                    if "No hits found" not in f.readline():
                        hits = []
                        i = 0
                        line = f.readline()
                        while i < 4 and line != "\n":
                            hits.append(line)
                            i += 1
                            line = f.readline()
                        if hit_with_species(hits, species) == -1 and line != "\n":
                            while line != "\n":
                                hits.append(line)
                                line = f.readline()
                            if hit_with_species(hits, species) != -1:
                                s.addRead(read_num, hits[0:hit_with_species(hits, species) + 1], length)
                            else:
                                s.addRead(read_num, hits[0:4], length)
                        else:
                            s.addRead(read_num, hits, length)
                elif line.startswith('>'):
                    hit_num = genomeid_in_hits(line, s, read_num)
                    if hit_num != 0:
                        while not line.startswith(" Identities"):
                            line = f.readline()
                        line = line.split()
                        frac = line[2]
                        percent = line[3][1:-2]
                        s.setFraction(read_num, hit_num, frac)
                        s.setPercent(read_num, hit_num, percent)
                        alignment = [f.readline(), f.readline()]
                        while not two_blank_lines(alignment):
                            alignment.append(f.readline())
                        sbjct = [line for line in alignment if line.startswith("Sbjct")]
                        cds_start = sbjct[0].split()[1]
                        last_row = sbjct[len(sbjct)-1]
                        cds_end = last_row.split()[3]
                        s.set_cds(read_num, hit_num, cds_start, cds_end)
                line = f.readline()
    print("Finished other columns for " + species)

"""Takes in a line from the BLAST output and the read number
for the sample. It finds the genome ID in the line and sees
if any of the hits in the read match. If they do, the hit number
(1-n) is returned, if not, 0 is returned."""
def genomeid_in_hits(line, sample, read_num):
    genomeid = line.split()[0]
    genomeid = genomeid[1:]
    for read in sample.reads:
        if read.n == read_num:
            sample_hits = read.blast_hits
            hit_num = 1
            for hit in sample_hits:
                genomeid2 = hit.line.split()[0]
                if genomeid == genomeid2:
                    return hit_num
                hit_num += 1
    return 0

"""Returns the first hit (0-n) that contains the species of interest,
or -1 if none of the hits contain it."""
def hit_with_species(list, species):
    for i in range(0, len(list)):
        if species in list[i]:
            return i
    return -1

"""Returns True if the last two lines are blank."""
def two_blank_lines(lines):
    if lines[len(lines)-2] == "\n" and lines[len(lines)-1] == "\n":
        return True
    return False

def write_to_excel(species, samples, sheet):
    sheet['A1'] = "Sample ID"
    sheet['B1'] = "Ct status"
    sheet['C1'] = "No. of " + species + " reads identified by Kraken"
    sheet['D1'] = "No. of reads with true identity to " + species
    sheet['E1'] = "First 4 Blast hits of confirmed " + species + " reads"
    sheet['F1'] = "Region (bp position)/CDS of " + species + " genome"
    sheet['G1'] = "Score (bits)"
    sheet['H1'] = "Supported e-value"
    sheet['I1'] = "% identity"
    sheet['J1'] = "Fraction identity"
    sheet['K1'] = "Read length (bp)"
    row = 2
    for s in samples:
        sheet.cell(row, 1).value = s.name
        if s.ct_positive:
            sheet.cell(row, 2).value = "Positive"
        else:
            sheet.cell(row, 2).value = "Negative"
        sheet.cell(row, 3).value = s.num_species_kraken
        print("Counting number of true hits for " + species)
        if species == "Chlamydia caviae":
            sheet.cell(row, 4).value = count_matches("Chlamydophila caviae", s)
        elif species == "Chlamydia unclassified":
            sheet.cell(row, 4).value = count_matches("Chlamydia sp.", s)
        else:
            sheet.cell(row, 4).value = count_matches(species, s)
        print("Finished counting number of true hits for " + species)
        for read in s.reads:
            sheet.cell(row, 5).value = "Read " + str(read.n)
            row += 1
            for hit in read.blast_hits:
                sheet.cell(row, 5).value = hit.line
                sheet.cell(row, 6).value = hit.cds_start + "-" + hit.cds_end
                sheet.cell(row, 7).value = hit.score
                sheet.cell(row, 8).value = hit.e_val
                sheet.cell(row, 9).value = hit.percent
                sheet.cell(row, 10).value = hit.fraction
                sheet.cell(row, 11).value = read.length
                if species == "Chlamydia caviae":
                    if "Chlamydophila caviae" in hit.line:
                        sheet.cell(row, 5).fill = PatternFill(start_color = "FFFF00", end_color = "FFFF00", fill_type = "solid")
                elif species == "Chlamydia unclassified":
                    if "Chlamydia sp." in hit.line:
                        sheet.cell(row, 5).fill = PatternFill(start_color = "FFFF00", end_color = "FFFF00", fill_type = "solid")
                else:
                    if species in hit.line:
                        sheet.cell(row, 5).fill = PatternFill(start_color = "FFFF00", end_color = "FFFF00", fill_type = "solid")
                row += 1

def main(species, blast_dir, sheet):
    samples = []
    samples.append(Sample("30C", False))
    samples.append(Sample("30R", False))
    samples.append(Sample("30V", False))
    samples.append(Sample("35C", False))
    samples.append(Sample("35R", False))
    samples.append(Sample("35V", False))
    samples.append(Sample("57C", False))
    samples.append(Sample("57R", False))
    samples.append(Sample("57V", False))
    samples.append(Sample("121C", False))
    samples.append(Sample("121R", False))
    samples.append(Sample("121V", False))
    samples.append(Sample("319C", False))
    samples.append(Sample("319R", False))
    samples.append(Sample("319V", False))
    samples.append(Sample("72C", True))
    samples.append(Sample("72R", True))
    samples.append(Sample("72V", True))
    samples.append(Sample("98C", True))
    samples.append(Sample("98R", True))
    samples.append(Sample("98V", True))
    samples.append(Sample("107C", True))
    samples.append(Sample("107R", True))
    samples.append(Sample("107V", True))
    samples.append(Sample("192C", True))
    samples.append(Sample("192R", True))
    samples.append(Sample("192V", True))
    samples.append(Sample("362C", True))
    samples.append(Sample("362R", True))
    samples.append(Sample("362V", True))
    count_kraken(species, samples)
    if species == "Chlamydia caviae":
        other_columns("Chlamydophila caviae", samples, blast_dir)
    elif species == "Chlamydia unclassified":
        other_columns("Chlamydia sp.", samples, blast_dir)
    else:
        other_columns(species, samples, blast_dir)
    write_to_excel(species, samples, sheet)

wb = Workbook()
ws = wb.active
ws.title = "Waddlia chondrophila"
ws2 = wb.create_sheet("Parachlamydia acanthamoebae")
ws3 = wb.create_sheet("Protochlamydia amoebophila")
ws4 = wb.create_sheet("Protochlamydia naegleriophila")
ws5 = wb.create_sheet("Simkania negevensis")
ws6 = wb.create_sheet("Neochlamydia")
main("Waddlia chondrophila", "/home/sbomman/ncbi_blast/Blasted_seq/Waddliaceae/waddlia/", ws)
main("Parachlamydia acanthamoebae", "/home/sbomman/ncbi_blast/Blasted_seq/Parachalmydiaceae/candidatus/acan/", ws2)
main("Protochlamydia amoebophila", "/home/sbomman/ncbi_blast/Blasted_seq/Parachalmydiaceae/candidatus/candidatus_protoch_amoebo/", ws3)
main("Protochlamydia naegleriophila", "/home/sbomman/ncbi_blast/Blasted_seq/Parachalmydiaceae/candidatus/candidatus_protoch_naegle/", ws4)
main("Simkania negevensis", "/home/sbomman/ncbi_blast/Blasted_seq/Simkania/", ws5)
main("Neochlamydia", "/home/sbomman/ncbi_blast/Blasted_seq/Neochlamydia_sp/", ws6)
wb.save("Blast_output_Chlamydia_Like_bacteria.xlsx")

wb2 = Workbook()
ws = wb2.active
ws.title = "Chlamydia abortus"
ws2 = wb2.create_sheet("Chlamydia avium")
ws3 = wb2.create_sheet("Chlamydia caviae")
ws4 = wb2.create_sheet("Chlamydia felis")
ws5 = wb2.create_sheet("Chlamydia gallinacea")
ws6 = wb2.create_sheet("Chlamydia muridarum")
ws7 = wb2.create_sheet("Chlamydia pecorum")
ws8 = wb2.create_sheet("Chlamydia pneumoniae")
ws9 = wb2.create_sheet("Chlamydia psittaci")
ws10 = wb2.create_sheet("Chlamydia suis")
ws11 = wb2.create_sheet("Chlamydia unclassified")
main("Chlamydia abortus", "/home/sbomman/ncbi_blast/Blasted_seq/Chlamydiaceae/Chlamydia/ch_abor/", ws)
main("Chlamydia avium", "/home/sbomman/ncbi_blast/Blasted_seq/Chlamydiaceae/Chlamydia/ch_aviu/", ws2)
main("Chlamydia caviae", "/home/sbomman/ncbi_blast/Blasted_seq/Chlamydiaceae/Chlamydia/ch_cavi/", ws3)
main("Chlamydia felis", "/home/sbomman/ncbi_blast/Blasted_seq/Chlamydiaceae/Chlamydia/ch_felis/", ws4)
main("Chlamydia gallinacea", "/home/sbomman/ncbi_blast/Blasted_seq/Chlamydiaceae/Chlamydia/ch_galli/", ws5)
main("Chlamydia muridarum", "/home/sbomman/ncbi_blast/Blasted_seq/Chlamydiaceae/Chlamydia/ch_muri/", ws6)
main("Chlamydia pecorum", "/home/sbomman/ncbi_blast/Blasted_seq/Chlamydiaceae/Chlamydia/ch_peco/", ws7)
main("Chlamydia pneumoniae", "/home/sbomman/ncbi_blast/Blasted_seq/Chlamydiaceae/Chlamydia/ch_pne/", ws8)
main("Chlamydia psittaci", "/home/sbomman/ncbi_blast/Blasted_seq/Chlamydiaceae/Chlamydia/ch_psit/", ws9)
main("Chlamydia suis", "/home/sbomman/ncbi_blast/Blasted_seq/Chlamydiaceae/Chlamydia/ch_suis/", ws10)
main("Chlamydia unclassified", "/home/sbomman/ncbi_blast/Blasted_seq/Chlamydiaceae/Chlamydia/ch_uncla/", ws11)
wb2.save("Blast_output_Chlamydia_species.xlsx")
