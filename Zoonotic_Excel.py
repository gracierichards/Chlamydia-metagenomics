import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell

status = {"30":"Negative", "35":"Negative", "57":"Negative", "121":"Negative", "319":"Negative", "72":"Positive", "98":"Positive",
"107":"Positive", "192":"Positive", "362":"Positive"}

class Sample:
    def __init__(self, name, ct_status):
        self.name = name
        self.ct_status = ct_status
        self.num_species_kraken = 0
        self.reads = []

    def addRead(self, n, blast_hits, length):
        self.reads.append(Read(n, blast_hits, length))

    def setFraction(self, read, hit, frac):
        for r in self.reads:
            if r.n == read:
                r.setFraction2(hit, frac)

    def setPercent(self, read, hit, percent):
        for r in self.reads:
            if r.n == read:
                r.setPercent2(hit, percent)

    def set_cds(self, read, hit, start, end):
        for r in self.reads:
            if r.n == read:
                r.set_cds2(hit, start, end)

class Read:
    def __init__(self, n, blast_hits, length):
        self.n = n
        self.length = length
        self.blast_hits = []
        for hit in blast_hits:
            line = hit[0:68]
            score = hit[70:74]
            e_val = hit[78:83]
            self.blast_hits.append(BlastHit(line, score, e_val))

    def setFraction2(self, hit, frac):
        self.blast_hits[hit-1].fraction = frac

    def setPercent2(self, hit, percent):
        self.blast_hits[hit-1].percent = percent

    def set_cds2(self, hit, start, end):
        self.blast_hits[hit-1].cds_start = start
        self.blast_hits[hit-1].cds_end = end

    def is_species(self, species):
        for hit in self.blast_hits:
            if species in hit.line:
                return True
        return False

class BlastHit:
    def __init__(self, line, score, e_val):
        self.line = line
        self.cds_start = ""
        self.cds_end = ""
        self.score = score
        self.e_val = e_val
        self.fraction = ""
        self.percent = ""

"""Searches through kraken reports to generate column C of the Excel file."""
def count_kraken(species, sample):
    print("Counting total number of reads for " + species)
    if species.endswith("includes all"):
        species.replace(" includes all", "")
    f = open("/home/sbomman/Kraken2/output/kraken_report_clean_" + sample.name + ".txt", "r")
    line = f.readline()
    while line != "":
        if species in line:
            words = line.split()
            sample.num_species_kraken = words[1]
            break
        line = f.readline()
    print("Finished counting total number of reads for " + species)

"""Calculates column D of the Excel file."""
def count_matches(species, sample):
    count = 0
    for read in sample.reads:
        if read.is_species(species):
            count += 1
    return count

"""Calculates all the other columns for a single sample object."""
def other_columns(species, s, path):
    print("Starting other columns for " + species)
    if species.endswith("includes all"):
        species.replace(" includes all", "")
    elif species.endswith("Turkey32"):
        species.replace(" str. Turkey32", "")
    elif species.endswith("Sterne"):
        species.replace(" str. Sterne", "")
    read_num = 0
    print("Sample " + s.name + ": " + str(s.num_species_kraken) + " reads")
    f = open(path, "r")
    line = f.readline()
    while line != "":
        if "Query=" in line:
            read_num += 1
            print("Read " + str(read_num))
            f.readline()
            line = f.readline()
            length = line[line.index('=')+1 :]
            f.readline()
            f.readline()
            if "No hits found" not in f.readline():
                hits = []
                i = 0
                line = f.readline()
                while i < 4 and line != "\n":
                    hits.append(line)
                    i += 1
                    line = f.readline()
                if hit_with_species(hits, species) == -1 and line != "\n":
                    while line != "\n":
                        hits.append(line)
                        line = f.readline()
                    if hit_with_species(hits, species) != -1:
                        s.addRead(read_num, hits[0:hit_with_species(hits, species) + 1], length)
                    else:
                        s.addRead(read_num, hits[0:4], length)
                else:
                    s.addRead(read_num, hits, length)
        elif line.startswith('>'):
            hit_num = genomeid_in_hits(line, s, read_num)
            if hit_num != 0:
                while not line.startswith(" Identities"):
                    line = f.readline()
                line = line.split()
                frac = line[2]
                percent = line[3][1:-2]
                s.setFraction(read_num, hit_num, frac)
                s.setPercent(read_num, hit_num, percent)
                alignment = [f.readline(), f.readline()]
                while not two_blank_lines(alignment):
                    alignment.append(f.readline())
                sbjct = [line for line in alignment if line.startswith("Sbjct")]
                cds_start = sbjct[0].split()[1]
                last_row = sbjct[len(sbjct)-1]
                cds_end = last_row.split()[3]
                s.set_cds(read_num, hit_num, cds_start, cds_end)
        line = f.readline()
    print("Finished other columns for " + species)

"""Takes in a line from the BLAST output and the read number
for the sample. It finds the genome ID in the line and sees
if any of the hits in the read match. If they do, the hit number
(1-n) is returned, if not, 0 is returned."""
def genomeid_in_hits(line, sample, read_num):
    genomeid = line.split()[0]
    genomeid = genomeid[1:]
    for read in sample.reads:
        if read.n == read_num:
            sample_hits = read.blast_hits
            hit_num = 1
            for hit in sample_hits:
                genomeid2 = hit.line.split()[0]
                if genomeid == genomeid2:
                    return hit_num
                hit_num += 1
    return 0

"""Returns the first hit (0-n) that contains the species of interest,
or -1 if none of the hits contain it."""
def hit_with_species(list, species):
    for i in range(0, len(list)):
        if species in list[i]:
            return i
    return -1

"""Returns True if the last two lines are blank."""
def two_blank_lines(lines):
    if lines[len(lines)-2] == "\n" and lines[len(lines)-1] == "\n":
        return True
    return False

def write_to_excel(species, samples, sheet):
    if species == "Bacillus anthracis includes all":
        sheet['A1'] = "This sheet/tab includes Bacillus anthracis, Bacillus anthracis str. CDC 684, Bacillus anthracis str. Sterne, and Bacillus anthracis str. Turkey32"
        sheet['A2'] = "Sample ID"
        sheet['B2'] = "Ct status"
        sheet['C2'] = "No. of " + species + " reads identified by Kraken"
        sheet['D2'] = "No. of reads with true identity to " + species
        sheet['E2'] = "First 4 Blast hits of confirmed " + species + " reads"
        sheet['F2'] = "Region (bp position)/CDS of " + species + " genome"
        sheet['G2'] = "Score (bits)"
        sheet['H2'] = "Supported e-value"
        sheet['I2'] = "% identity"
        sheet['J2'] = "Fraction identity"
        sheet['K2'] = "Read length (bp)"
        row = 3
    else:
        sheet['A1'] = "Sample ID"
        sheet['B1'] = "Ct status"
        sheet['C1'] = "No. of " + species + " reads identified by Kraken"
        sheet['D1'] = "No. of reads with true identity to " + species
        sheet['E1'] = "First 4 Blast hits of confirmed " + species + " reads"
        sheet['F1'] = "Region (bp position)/CDS of " + species + " genome"
        sheet['G1'] = "Score (bits)"
        sheet['H1'] = "Supported e-value"
        sheet['I1'] = "% identity"
        sheet['J1'] = "Fraction identity"
        sheet['K1'] = "Read length (bp)"
        row = 2
    for s in samples:
        sheet.cell(row, 1).value = s.name
        sheet.cell(row, 2).value = s.ct_status
        sheet.cell(row, 3).value = s.num_species_kraken
        print("Counting number of true hits for " + species)
        if species.endswith("includes all"):
            species.replace(" includes all", "")
        elif species.endswith("Turkey32"):
            species.replace(" str. Turkey32", "")
        elif species.endswith("Sterne"):
            species.replace(" str. Sterne", "")
        sheet.cell(row, 4).value = count_matches(species, s)
        print("Finished counting number of true hits for " + species)
        for read in s.reads:
            sheet.cell(row, 5).value = "Read " + str(read.n)
            row += 1
            sheet.cell(row, 1).value = s.name
            for hit in read.blast_hits:
                sheet.cell(row, 5).value = hit.line
                sheet.cell(row, 6).value = hit.cds_start + "-" + hit.cds_end
                sheet.cell(row, 7).value = hit.score
                sheet.cell(row, 8).value = hit.e_val
                sheet.cell(row, 9).value = hit.percent
                sheet.cell(row, 10).value = hit.fraction
                sheet.cell(row, 11).value = read.length
                if species in hit.line:
                    sheet.cell(row, 5).fill = PatternFill(start_color = "FFFF00", end_color = "FFFF00", fill_type = "solid")
                row += 1
                sheet.cell(row, 1).value = s.name

def collect_species(dir):
    s = set()
    for f in os.listdir(dir):
        if "_" in f:
            first_score = f.index("_")
            f = f[first_score+1 : -4]
            s.add(f)
    return s

for dir in os.listdir("/home/sbomman/ncbi_blast/Blasted_seq/Zoonotic_pathogens"):
    if not dir.endswith("txt"):
        wb = Workbook()
        dire = "/home/sbomman/ncbi_blast/Blasted_seq/Zoonotic_pathogens/" + dir + "/"
        species_list = collect_species(dire)
        for species in species_list:
            species = species.replace("_", " ")
            ws = wb.create_sheet(species[-31:])
            samples = []
            for file in os.listdir(dire):
                first_score = file.index("_")
                file_wspace = file[first_score+1 : -4]
                file_wspace = file_wspace.replace("_", " ")
                if file_wspace == species:
                    sample_name = file[0:first_score]
                    sample_num = sample_name[0:-1]
                    s = Sample(sample_name, status[sample_num])
                    count_kraken(species, s)
                    other_columns(species, s, dire + file)
                    samples.append(s)
            write_to_excel(species, samples, ws)
        wb.save("Blast_output_" + dir + ".xlsx")
