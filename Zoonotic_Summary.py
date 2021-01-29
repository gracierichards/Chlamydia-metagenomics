import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell

samples = ["30C", "30R", "30V", "35C", "35R", "35V", "57C", "57R", "57V", "121C", "121R", "121V", "319C", "319R", "319V", "72C", "72R", "72V",
"98C", "98R", "98V", "107C", "107R", "107V", "192C", "192R", "192V", "362C", "362R", "362V"]

class Sample:
    def __init__(self, name):
        self.name = name
        self.num_species_kraken = 0
        self.reads = []

    def addRead(self, n, blast_hits, length):
        self.reads.append(Read(n, blast_hits, length))

    def setFraction(self, read, hit, frac):
        for r in self.reads:
            if r.n == read:
                r.setFraction2(hit, frac)

    def setPercent(self, read, hit, percent):
        for r in self.reads:
            if r.n == read:
                r.setPercent2(hit, percent)

    def set_cds(self, read, hit, start, end):
        for r in self.reads:
            if r.n == read:
                r.set_cds2(hit, start, end)

class Read:
    def __init__(self, n, blast_hits, length):
        self.n = n
        self.length = length
        self.blast_hits = []
        for hit in blast_hits:
            line = hit[0:68]
            score = hit[70:74]
            e_val = hit[78:83]
            self.blast_hits.append(BlastHit(line, score, e_val))

    def setFraction2(self, hit, frac):
        self.blast_hits[hit-1].fraction = frac

    def setPercent2(self, hit, percent):
        self.blast_hits[hit-1].percent = percent

    def set_cds2(self, hit, start, end):
        self.blast_hits[hit-1].cds_start = start
        self.blast_hits[hit-1].cds_end = end

    def is_species(self, species):
        for hit in self.blast_hits:
            if species in hit.line:
                return True
        return False

class BlastHit:
    def __init__(self, line, score, e_val):
        self.line = line
        self.cds_start = ""
        self.cds_end = ""
        self.score = score
        self.e_val = e_val
        self.fraction = ""
        self.percent = ""

"""Searches through kraken reports to generate column C of the Excel file."""
def count_kraken(species, sample):
    print("Counting total number of reads for " + species)
    if species.endswith("includes all"):
        species.replace(" includes all", "")
    f = open("/home/sbomman/Kraken2/output/kraken_report_clean_" + sample.name + ".txt", "r")
    line = f.readline()
    while line != "":
        if species in line:
            words = line.split()
            sample.num_species_kraken = words[1]
            break
        line = f.readline()
    print("Finished counting total number of reads for " + species)

"""Calculates column D of the Excel file."""
def count_matches(species, sample):
    count = 0
    for read in sample.reads:
        if read.is_species(species):
            count += 1
    return count

"""Calculates all the other columns for a single sample object."""
def other_columns(species, s, path):
    print("Starting other columns for " + species)
    read_num = 0
    print("Sample " + s.name + ": " + str(s.num_species_kraken) + " reads")
    f = open(path, "r")
    line = f.readline()
    while line != "":
        if "Query=" in line:
            read_num += 1
            print("Read " + str(read_num))
            f.readline()
            line = f.readline()
            length = line[line.index('=')+1 :]
            f.readline()
            f.readline()
            if "No hits found" not in f.readline():
                hits = []
                i = 0
                line = f.readline()
                while i < 4 and line != "\n":
                    hits.append(line)
                    i += 1
                    line = f.readline()
                if hit_with_species(hits, species) == -1 and line != "\n":
                    while line != "\n":
                        hits.append(line)
                        line = f.readline()
                    if hit_with_species(hits, species) != -1:
                        s.addRead(read_num, hits[0:hit_with_species(hits, species) + 1], length)
                    else:
                        s.addRead(read_num, hits[0:4], length)
                else:
                    s.addRead(read_num, hits, length)
        line = f.readline()
    print("Finished other columns for " + species)

"""Takes in a line from the BLAST output and the read number
for the sample. It finds the genome ID in the line and sees
if any of the hits in the read match. If they do, the hit number
(1-n) is returned, if not, 0 is returned."""
def genomeid_in_hits(line, sample, read_num):
    genomeid = line.split()[0]
    genomeid = genomeid[1:]
    for read in sample.reads:
        if read.n == read_num:
            sample_hits = read.blast_hits
            hit_num = 1
            for hit in sample_hits:
                genomeid2 = hit.line.split()[0]
                if genomeid == genomeid2:
                    return hit_num
                hit_num += 1
    return 0

"""Returns the first hit (0-n) that contains the species of interest,
or -1 if none of the hits contain it."""
def hit_with_species(list, species):
    for i in range(0, len(list)):
        if species in list[i]:
            return i
    return -1

"""Returns True if the last two lines are blank."""
def two_blank_lines(lines):
    if lines[len(lines)-2] == "\n" and lines[len(lines)-1] == "\n":
        return True
    return False

"""Creates a list of all species in the directory with no duplicates."""
def collect_species(dir):
    s = set()
    for f in os.listdir(dir):
        if "_" in f:
            first_score = f.index("_")
            f = f[first_score+1 : -4]
            s.add(f)
    return s

"""Finds which row to put the data in."""
def find_row(sheet, sample_name):
    return samples.index(sample_name) + 2

wb = Workbook()
ws = wb.active
ws.cell(1, 1).value = "Sample"
row = 2
for s in samples:
    ws.cell(row, 1).value = s
    row += 1

column = 2
for dir in os.listdir("/home/sbomman/ncbi_blast/Blasted_seq/Zoonotic_pathogens"):
    if not dir.endswith("txt") and not dir.endswith("py"):
        dire = "/home/sbomman/ncbi_blast/Blasted_seq/Zoonotic_pathogens/" + dir + "/"
        species_list = collect_species(dire)
        for species in species_list:
            species = species.replace("_", " ")
            ws.cell(1, column).value = species
            if species.endswith("includes all"):
                species.replace(" includes all", "")
            elif species.endswith("Turkey32"):
                species.replace(" str. Turkey32", "")
            elif species.endswith("Sterne"):
                species.replace(" str. Sterne", "")
            for file in os.listdir(dire):
                if file.endswith("txt"):
                    first_score = file.index("_")
                    file_wspace = file[first_score+1 : -4]
                    file_wspace = file_wspace.replace("_", " ")
                    if file_wspace == species:
                        sample_name = file[0:first_score]
                        s = Sample(sample_name)
                        count_kraken(species, s)
                        other_columns(species, s, dire + file)
                        numerator = count_matches(species, s)
                        denom = s.num_species_kraken
                        row = find_row(ws, sample_name)
                        ws.cell(row, column).value = str(numerator) + "/" + str(denom)
            column += 1
        wb.save("Blast_output_" + dir + ".xlsx")
