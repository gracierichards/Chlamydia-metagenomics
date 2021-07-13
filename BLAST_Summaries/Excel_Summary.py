#import os
from os import path
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell

class Sample:
    def __init__(self, name, ct_positive):
        self.name = name
        self.ct_positive = ct_positive
        self.num_species_kraken = 0
        self.reads = []

    def addRead(self, n, blast_hits):
        self.reads.append(Read(n, blast_hits))

    def setPercent(self, read, hit, percent):
        for r in self.reads:
            if r.n == read:
                r.setPercent2(hit, percent)

    def set_cds(self, read, hit, start, end):
        for r in self.reads:
            if r.n == read:
                r.set_cds2(hit, start, end)

class Read:
    def __init__(self, n, blast_hits):
        self.n = n
        self.blast_hits = []
        for hit in blast_hits:
            line = hit[0:68]
            score = hit[70:74]
            e_val = hit[78:83]
            self.blast_hits.append(BlastHit(line, score, e_val))

    def setPercent2(self, hit, percent):
        self.blast_hits[hit-1].percent = percent

    def set_cds2(self, hit, start, end):
        self.blast_hits[hit-1].cds_start = start
        self.blast_hits[hit-1].cds_end = end

    def is_species(self, species):
        for hit in self.blast_hits:
            if species in hit.line:
                return True
        return False

class BlastHit:
    def __init__(self, line, score, e_val):
        self.line = line
        self.cds_start = ""
        self.cds_end = ""
        self.score = score
        self.e_val = e_val
        self.percent = ""

"""Searches through kraken reports to generate column C of the Excel file."""
def count_kraken(species, samples):
    for s in samples:
        f = open("/home/sbomman/Kraken2/output/kraken_report_clean_" + s.name + ".txt", "r")
        line = f.readline()
        while line != "":
            if species in line:
                words = line.split()
                s.num_species_kraken = words[1]
                break
            line = f.readline()

"""Calculates column D of the Excel file."""
def count_matches(species, sample):
    count = 0
    for read in sample.reads:
        if read.is_species(species):
            count += 1
    return count

"""Calculates all the other columns and stores the data in the sample objects."""
def other_columns(species, samples, blast_dir):
    for s in samples:
        read_num = 0
        filename = blast_dir + s.name + "_blasted.txt"
        if path.exists(filename):
            f = open(filename, "r")
            line = f.readline()
            while line != "":
                if "Query=" in line:
                    read_num += 1
                    for i in range(0, 4):
                        f.readline()
                    if "No hits found" not in f.readline():
                        hits = []
                        i = 0
                        line = f.readline()
                        while i < 4 and line != "\n":
                            hits.append(line)
                            i += 1
                            line = f.readline()
                        if hit_with_species(hits, species) == -1 and line != "\n":
                            while line != "\n":
                                hits.append(line)
                                line = f.readline()
                            if hit_with_species(hits, species) != -1:
                                s.addRead(read_num, hits[0:hit_with_species(hits, species) + 1])
                            else:
                                s.addRead(read_num, hits[0:4])
                        else:
                            s.addRead(read_num, hits)
                line = f.readline()


"""Takes in a line from the BLAST output and the read number
for the sample. It finds the genome ID in the line and sees
if any of the hits in the read match. If they do, the hit number
(1-n) is returned, if not, 0 is returned."""
def genomeid_in_hits(line, sample, read_num):
    genomeid = line.split()[0]
    genomeid = genomeid[1:]
    for read in sample.reads:
        if read.n == read_num:
            sample_hits = read.blast_hits
            hit_num = 1
            for hit in sample_hits:
                genomeid2 = hit.line.split()[0]
                if genomeid == genomeid2:
                    return hit_num
                hit_num += 1
    return 0

"""Returns the first hit (0-n) that contains the species of interest,
or -1 if none of the hits contain it."""
def hit_with_species(list, species):
    for i in range(0, len(list)):
        if species in list[i]:
            return i
    return -1

"""Returns True if the last two lines are blank."""
def two_blank_lines(lines):
    if lines[len(lines)-2] == "\n" and lines[len(lines)-1] == "\n":
        return True
    return False

def write_to_excel(species, samples, column):
    row = 2
    for s in samples:
        numerator = count_matches(species, s)
        denom = s.num_species_kraken
        ws.cell(row, column).value = str(numerator) + "/" + str(denom)
        row += 1

def main(species, blast_dir, column):
    samples = []
    samples.append(Sample("30C", False))
    samples.append(Sample("30R", False))
    samples.append(Sample("30V", False))
    samples.append(Sample("35C", False))
    samples.append(Sample("35R", False))
    samples.append(Sample("35V", False))
    samples.append(Sample("57C", False))
    samples.append(Sample("57R", False))
    samples.append(Sample("57V", False))
    samples.append(Sample("121C", False))
    samples.append(Sample("121R", False))
    samples.append(Sample("121V", False))
    samples.append(Sample("319C", False))
    samples.append(Sample("319R", False))
    samples.append(Sample("319V", False))
    samples.append(Sample("72C", True))
    samples.append(Sample("72R", True))
    samples.append(Sample("72V", True))
    samples.append(Sample("98C", True))
    samples.append(Sample("98R", True))
    samples.append(Sample("98V", True))
    samples.append(Sample("107C", True))
    samples.append(Sample("107R", True))
    samples.append(Sample("107V", True))
    samples.append(Sample("192C", True))
    samples.append(Sample("192R", True))
    samples.append(Sample("192V", True))
    samples.append(Sample("362C", True))
    samples.append(Sample("362R", True))
    samples.append(Sample("362V", True))
    if species == "Chlamydia trachomatis":
        row = 2
        for s in samples:
            ws.cell(row, 1).value = s.name
            row += 1
    count_kraken(species, samples)
    if species == "Chlamydia caviae":
        other_columns("Chlamydophila caviae", samples, blast_dir)
        write_to_excel("Chlamydophila caviae", samples, column)
    elif species == "unclassified Chlamydia":
        other_columns("Chlamydia sp.", samples, blast_dir)
        write_to_excel("Chlamydia sp.", samples, column)
    else:
        other_columns(species, samples, blast_dir)
        write_to_excel(species, samples, column)

wb = Workbook()
ws = wb.active
ws.cell(1, 1).value = "Sample"
ws.cell(1, 2).value = "Chlamydia trachomatis"
ws.cell(1, 3).value = "Chlamydia abortus"
ws.cell(1, 4).value = "Chlamydia avium"
ws.cell(1, 5).value = "Chlamydia caviae"
ws.cell(1, 6).value = "Chlamydia felis"
ws.cell(1, 7).value = "Chlamydia gallinacea"
ws.cell(1, 8).value = "Chlamydia muridarum"
ws.cell(1, 9).value = "Chlamydia pecorum"
ws.cell(1, 10).value = "Chlamydia pneumoniae"
ws.cell(1, 11).value = "Chlamydia psittaci"
ws.cell(1, 12).value = "Chlamydia suis"
ws.cell(1, 13).value = "unclassified Chlamydia"
ws.cell(1, 14).value = "Waddlia chondrophila"
ws.cell(1, 15).value = "Parachlamydia acanthamoebae"
ws.cell(1, 16).value = "Protochlamydia amoebophila"
ws.cell(1, 17).value = "Protochlamydia naegleriophila"
ws.cell(1, 18).value = "Simkania negevensis"
ws.cell(1, 19).value = "Neochlamydia"
main("Chlamydia trachomatis", "/home/sbomman/ncbi_blast/Blasted_seq/Chlamydiaceae/Chlamydia/ch_tra/", 2)
main("Chlamydia abortus", "/home/sbomman/ncbi_blast/Blasted_seq/Chlamydiaceae/Chlamydia/ch_abor/", 3)
main("Chlamydia avium", "/home/sbomman/ncbi_blast/Blasted_seq/Chlamydiaceae/Chlamydia/ch_aviu/", 4)
main("Chlamydia caviae", "/home/sbomman/ncbi_blast/Blasted_seq/Chlamydiaceae/Chlamydia/ch_cavi/", 5)
main("Chlamydia felis", "/home/sbomman/ncbi_blast/Blasted_seq/Chlamydiaceae/Chlamydia/ch_felis/", 6)
main("Chlamydia gallinacea", "/home/sbomman/ncbi_blast/Blasted_seq/Chlamydiaceae/Chlamydia/ch_galli/", 7)
main("Chlamydia muridarum", "/home/sbomman/ncbi_blast/Blasted_seq/Chlamydiaceae/Chlamydia/ch_muri/", 8)
main("Chlamydia pecorum", "/home/sbomman/ncbi_blast/Blasted_seq/Chlamydiaceae/Chlamydia/ch_peco/", 9)
main("Chlamydia pneumoniae", "/home/sbomman/ncbi_blast/Blasted_seq/Chlamydiaceae/Chlamydia/ch_pne/", 10)
main("Chlamydia psittaci", "/home/sbomman/ncbi_blast/Blasted_seq/Chlamydiaceae/Chlamydia/ch_psit/", 11)
main("Chlamydia suis", "/home/sbomman/ncbi_blast/Blasted_seq/Chlamydiaceae/Chlamydia/ch_suis/", 12)
main("unclassified Chlamydia", "/home/sbomman/ncbi_blast/Blasted_seq/Chlamydiaceae/Chlamydia/ch_uncla/", 13)
main("Waddlia chondrophila", "/home/sbomman/ncbi_blast/Blasted_seq/Waddliaceae/waddlia/", 14)
main("Parachlamydia acanthamoebae", "/home/sbomman/ncbi_blast/Blasted_seq/Parachalmydiaceae/candidatus/acan/", 15)
main("Protochlamydia amoebophila", "/home/sbomman/ncbi_blast/Blasted_seq/Parachalmydiaceae/candidatus/candidatus_protoch_amoebo/", 16)
main("Protochlamydia naegleriophila", "/home/sbomman/ncbi_blast/Blasted_seq/Parachalmydiaceae/candidatus/candidatus_protoch_naegle/", 17)
main("Simkania negevensis", "/home/sbomman/ncbi_blast/Blasted_seq/Simkania/", 18)
main("Neochlamydia", "/home/sbomman/ncbi_blast/Blasted_seq/Neochlamydia_sp/", 19)
wb.save("Blast_Summary.xlsx")
