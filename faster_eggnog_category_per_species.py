import os
import openpyxl
from openpyxl import Workbook

# Add the files that you know for sure have finished running here, then this program won't include them.
# completed = ["1016_ds.##########.fq.gz.EggNOG.annotation.txt", "1197TXMT_ds.##########.fq.gz.EggNOG.annotation.txt", "1040C_virgo.EggNOG.annotation.txt"]
completed = []

categories = ["J", "A", "K", "L", "B", "D", "Y", "V", "T", "M", "N", "Z", "W", "U", "O", "C", "G", "E", "F", "H", "I", "P", "Q", "R", "S"]

species_table_hashset = {}
species_table = open("/home/sbomman/VIRGO/downloads.igs.umaryland.edu/virgo/VIRGO/1_VIRGO/1.taxon.tbl.txt", "r")
line = species_table.readline()
while line != "":
    gene = line.split()[1]
    species0 = line.split()[2]
    species_table_hashset[gene] = species0
    line = species_table.readline()

class Species:
    def __init__(self, name):
        self.name = name
        self.categories = {"J": 0, "A":0, "K": 0, "L":0, "B":0, "D":0, "Y":0, "V":0, "T": 0, "M":0, "N":0, "Z":0, "W":0,
                           "U":0, "O":0, "C":0,"G":0, "E":0, "F":0, "H":0, "I":0, "P":0, "Q":0, "R":0,"S":0}
    
def in_species_list(s0, species_list):
    for s1 in species_list:
        if s1.name == s0:
            return True
    return False

def find_species(s0, species_list):
    for s1 in species_list:
        if s1.name == s0:
            return s1

# The path to the directory containing your EggNOG files of interest
for path0 in os.listdir("temp_mapping"):
    if "EggNOG" in path0:
        if path0 not in completed:
            eggnog = open("temp_mapping/" + path0, 'r')
            wb = Workbook()
            ws = wb.active
            row = 2
            col = 3
            for c in categories:
                ws.cell(1, col).value = c
                col += 1
            species_list = []
            line = eggnog.readline()
            line = eggnog.readline()
            while line != "":
                gene = line.split('\t')[0]
                category = line.split('\t')[1]
                if gene in species_table_hashset:
                    species_str = species_table_hashset[gene]
                else:
                    species_str = "Hypothetical protein"
                if in_species_list(species_str, species_list):
                    species_obj = find_species(species_str, species_list)
                    if ',' in category:
                        letters = category.split(", ")
                        species_obj.categories[letters[0]] = species_obj.categories[letters[0]] + 1
                        species_obj.categories[letters[1]] = species_obj.categories[letters[1]] + 1
                    else:
                        species_obj.categories[category] = species_obj.categories[category] + 1
                else:
                    species_obj = Species(species_str)
                    if ',' in category:
                        letters = category.split(", ")
                        species_obj.categories[letters[0]] = 1
                        species_obj.categories[letters[1]] = 1
                    else:
                        species_obj.categories[category] = 1
                    species_list.append(species_obj)
                line = eggnog.readline()
            """Find the sample name, which is just a substring of the file name"""
            if "_" in path0:
                sample = path0[0:path0.index("_")]
            elif ".fq.gz" in path0:
                sample = path0[0:path0.index(".fq.gz")]
            else:
                print("Format of file name not recognized. Skipping file:", path0)
            ws.cell(row, 1).value = sample
            for s in species_list:
                if s.name != "Hypothetical protein":
                    ws.cell(row, 2).value = s.name
                    col = 3
                    for c in categories:
                        ws.cell(row, col).value = s.categories[c]
                        col += 1
                    row += 1

            wb.save("eggnog_" + sample + ".xlsx")