import os
import openpyxl

"""Remember to upload Excel file to server, change path if needed"""
wb = openpyxl.load_workbook(filename = "/home/sbomman/mbio_updated.xlsx")
ws = wb["Bactopia_Chlamydia_Trachomatis"]
sample = ""
r = 1
dict0 = {}
"""Finds how many rows of samples before 26V there are"""
while sample != "26V":
    sample = ws.cell(r, 1).value
    r += 1
"""Making the dictionary"""
#print(r)
for r2 in range(2, r):
    sample = ws.cell(r2, 1).value
    sample = sample.replace("/", "_")
    accession = ws.cell(r2, 3).value
    if accession != None:
        #print(accession)
        accession = "".join(accession.split())
        dict0[accession] = sample
    r2 += 1
"""Recursively renames all accession numbers found in the path, path+ending"""
def rename_all(path, ending):
    #print("Start of function, path is " + path + ", ending is " + ending)
    if ending.startswith("ERX"):
        #print("Ending starts with ERX")
        dash = 100
        period = 100
        underscore = 100
        if "-" in ending:
            dash = ending.index("-")
        if "." in ending:
            period = ending.index(".")
        if "_" in ending:
            underscore = ending.index("_")
        accession = ending[0 : min(dash, period, underscore, len(ending))]
        new_name = ending.replace(accession, dict0[accession])
        print("Renaming " + path + ending + " to " + path + new_name)
        os.rename(path + ending, path + new_name)
        path = path + new_name + "/"
    else:
        path = path + ending + "/"
    #print("path is " + path)
    if not os.path.isdir(path):
        #print("There are no subdirectories")
        return
    #print("There are subdirectories")
    for i in os.listdir(path):
        rename_all(path, i)
        
"""Change this: this is the starting directory. Break it up into (the rest of the path), (everything after the last /)"""
rename_all("/home/sbomman/", "bactopia_new")
        